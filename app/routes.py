from flask import Blueprint, request, jsonify, render_template, current_app, send_file, Response
from models import (
    db, Task, Executor, FileDocument, YandexReport, YandexReportTaskLink,
    NotificationSettings, NotificationRecipient, TaskNotificationLog,
)
from pdf_parser.pdf_engine import parse_pdf
from portal_auth import require_write, require_admin, require_read, get_current_user
import hashlib
import io
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timedelta
from sqlalchemy import delete, update

main = Blueprint('main', __name__)
KCHZ_EXCEL_FILENAME = "kchz_traffic_light.xlsx"


def _excel_storage_path():
    storage_dir = os.path.join(current_app.root_path, "uploads")
    os.makedirs(storage_dir, exist_ok=True)
    return os.path.join(storage_dir, KCHZ_EXCEL_FILENAME)


def _safe_deadline(task):
    # В parser используются "дальние" даты как маркер отсутствия срока.
    if not task.deadline:
        return None
    if task.deadline.year >= 2099:
        return None
    return task.deadline


def _deadline_flag(task, enabled=True):
    if not enabled or not task.deadline or task.deadline.year >= 2099 or task.status == "Выполнено":
        return "normal"
    days_left = (task.deadline.date() - datetime.now().date()).days
    if days_left < 0:
        return "overdue"
    if days_left <= 3:
        return "due3"
    if days_left <= 7:
        return "due7"
    return "normal"


def _clean_mass_title(title, item_number):
    text = (title or "").strip()
    if not text:
        return f"Пункт {item_number}" if item_number else "Поручение"
    lowered = text.lower()
    noisy_phrases = [
        "органам местного самоуправления республики саха",
        "главам муниципальных образований республики саха",
        "рекомендовать органам местного самоуправления",
    ]
    if any(phrase in lowered for phrase in noisy_phrases):
        return f"Пункт {item_number}" if item_number else "Поручение для ОМСУ"
    text = re.sub(r'^\s*пункт\s*\d+(?:\.\d+)*\.?\s*', '', text, flags=re.IGNORECASE).strip()
    if item_number:
        return f"Пункт {item_number}. {text}" if text else f"Пункт {item_number}"
    return text or "Поручение"


def _mass_group_key(task):
    deadline_str = task.deadline.strftime('%Y-%m-%d') if task.deadline else ''
    return (task.item_number or '', task.title or '', deadline_str)


def _mass_task_view_stats(group_tasks):
    """Агрегат для массового поручения: одна строка, счётчик — оставшиеся районы."""
    total = len(group_tasks)
    done = sum(1 for t in group_tasks if _task_verified_by_disk(t))
    remaining = total - done
    open_tasks = [t for t in group_tasks if not _task_verified_by_disk(t)]
    now = datetime.now()

    has_overdue = any(
        t.status == 'Просрочено'
        or (_safe_deadline(t) and _safe_deadline(t) < now)
        for t in open_tasks
    )
    if remaining == 0 and done > 0:
        display_status = 'Выполнено'
    elif has_overdue:
        display_status = 'Просрочено'
    else:
        display_status = 'В работе'

    rep = min(group_tasks, key=lambda x: x.id)
    setattr(rep, 'group_executor', 'Органы местного самоуправления Республики Саха (Якутия)')
    setattr(rep, 'mass_view', True)
    setattr(rep, 'mass_count', remaining)
    setattr(rep, 'mass_count_total', total)
    setattr(rep, 'mass_count_done', done)
    setattr(rep, 'mass_all_done', remaining == 0 and done > 0)
    setattr(rep, 'mass_title', _clean_mass_title(rep.title, rep.item_number))
    setattr(rep, 'display_status', display_status)
    return rep


def _get_district_names():
    return set(_get_district_links().keys())


def _latest_report_for_task(task):
    report = YandexReport.query.filter_by(task_id=task.id).order_by(
        YandexReport.received_at.desc()
    ).first()
    if report:
        return report
    link = (
        YandexReportTaskLink.query.filter_by(task_id=task.id)
        .join(YandexReport)
        .order_by(YandexReport.received_at.desc())
        .first()
    )
    return link.report if link else None


def _task_verified_by_disk(task):
    """Поручение считается выполненным только при отчёте с Яндекс.Диска."""
    return _latest_report_for_task(task) is not None


def _get_district_links():
    try:
        from services.districts import DISTRICTS
        return DISTRICTS
    except Exception:
        return {}


def _task_timing_status(task):
    deadline = _safe_deadline(task)
    report = _latest_report_for_task(task)

    if not _task_verified_by_disk(task):
        if deadline and deadline < datetime.now():
            return "overdue_open"
        return "open"

    if not deadline:
        return "done_no_deadline"
    if report and report.received_at and report.received_at <= deadline:
        return "done_in_time"
    if report and report.received_at and report.received_at > deadline:
        return "done_late"
    return "done_no_date"


def _build_leaderboard_data():
    executors = Executor.query.all()
    leaders = []
    now = datetime.now()
    district_names = _get_district_names()
    district_links = _get_district_links()

    for ex in executors:
        if district_names and ex.name not in district_names:
            continue
        tasks = Task.query.filter_by(executor_id=ex.id).order_by(Task.deadline.asc()).all()
        if not tasks:
            continue

        verified_completed = sum(1 for t in tasks if _task_verified_by_disk(t))
        manual_without_disk = sum(
            1 for t in tasks
            if t.status == "Выполнено" and not _task_verified_by_disk(t)
        )
        overdue = sum(
            1 for t in tasks
            if not _task_verified_by_disk(t) and _safe_deadline(t) and _safe_deadline(t) < now
        )
        in_progress = len(tasks) - verified_completed - overdue
        percentage = int((verified_completed / len(tasks)) * 100) if tasks else 0

        timing = {
            "done_in_time": 0,
            "done_late": 0,
            "done_no_deadline": 0,
            "done_no_date": 0,
            "open": 0,
            "overdue_open": 0,
        }
        for task in tasks:
            timing[_task_timing_status(task)] += 1
            report = _latest_report_for_task(task)
            setattr(task, 'timing_status', _task_timing_status(task))
            setattr(task, 'report_date', report.received_at if report else None)
            setattr(task, 'verified_by_disk', _task_verified_by_disk(task))

        if percentage >= 90 and overdue == 0 and verified_completed > 0:
            color = "green"
        elif percentage >= 70:
            color = "yellow"
        else:
            color = "red"

        last_report = YandexReport.query.filter_by(executor_id=ex.id).order_by(YandexReport.received_at.desc()).first()
        leaders.append({
            "id": ex.id,
            "name": ex.name,
            "total": len(tasks),
            "completed": verified_completed,
            "manual_without_disk": manual_without_disk,
            "overdue": overdue,
            "in_progress": in_progress,
            "percentage": percentage,
            "color": color,
            "district_url": district_links.get(ex.name),
            "reports_count": YandexReport.query.filter_by(executor_id=ex.id).count(),
            "last_report_date": last_report.received_at.strftime("%d.%m.%Y %H:%M") if last_report and last_report.received_at else None,
            "last_report_file": last_report.filename if last_report else None,
            "timing": timing,
            "tasks": tasks,
        })

    color_order = {"green": 0, "yellow": 1, "red": 2, "gray": 3}
    leaders.sort(key=lambda x: (color_order.get(x["color"], 3), -x["percentage"], x["name"]))

    traffic_stats = {
        "total": len(leaders),
        "green": sum(1 for l in leaders if l["color"] == "green"),
        "yellow": sum(1 for l in leaders if l["color"] == "yellow"),
        "red": sum(1 for l in leaders if l["color"] == "red"),
    }
    return leaders, traffic_stats


def _period_bounds(period_key):
    now = datetime.now()
    if period_key == "3m":
        return now - timedelta(days=90), now
    if period_key == "6m":
        return now - timedelta(days=180), now
    if period_key == "1y":
        return now - timedelta(days=365), now
    return now - timedelta(days=30), now


def _favorite_metrics(executor, period_key):
    start_date, end_date = _period_bounds(period_key)
    q = Task.query.filter(
        Task.executor_id == executor.id,
        Task.created_at >= start_date,
        Task.created_at <= end_date
    ).order_by(Task.created_at.asc())
    tasks = q.all()

    in_time = 0
    late = 0
    no_deadline = 0
    open_count = 0
    for task in tasks:
        status = _task_timing_status(task)
        if status == "done_in_time":
            in_time += 1
        elif status in ("done_late", "done_no_date"):
            late += 1
        elif status == "done_no_deadline":
            no_deadline += 1
        else:
            open_count += 1

    total = len(tasks)
    completed = in_time + late + no_deadline
    performance_base = in_time + late
    performance_pct = int((in_time / performance_base) * 100) if performance_base > 0 else 0
    overall_pct = int((completed / total) * 100) if total > 0 else 0

    if total == 0:
        color = "gray"
    elif performance_pct >= 90 and late == 0:
        color = "green"
    elif performance_pct >= 70:
        color = "yellow"
    else:
        color = "red"

    return {
        "executor": executor.name,
        "period": period_key,
        "from": start_date.strftime("%d.%m.%Y"),
        "to": end_date.strftime("%d.%m.%Y"),
        "total": total,
        "completed": completed,
        "in_time": in_time,
        "late": late,
        "no_deadline": no_deadline,
        "open": open_count,
        "performance_pct": performance_pct,
        "overall_pct": overall_pct,
        "traffic_color": color,
    }, tasks

# ─── Страницы ───────────────────────────────────────────────────────────

@main.route('/')
def index():
    return render_template('index.html')

@main.route('/dashboard')
def dashboard():
    search_q = request.args.get('q', '').strip()
    debug_mode = request.args.get('debug', '0') == '1'
    tasks = Task.query.all()

    # Calculate stats for the dashboard cards
    total_tasks = len(tasks)
    completed_tasks = Task.query.filter_by(status='Выполнено').count()
    overdue_tasks = Task.query.filter_by(status='Просрочено').count()
    in_progress_tasks = Task.query.filter(Task.status.in_(['В работе', 'Ожидает'])).count()

    percentage = int((completed_tasks / total_tasks * 100)) if total_tasks > 0 else 0

    now = datetime.now()
    current_year = now.year
    year_docs = FileDocument.query.filter(FileDocument.uploaded_at >= datetime(current_year, 1, 1)).count()
    year_tasks = Task.query.filter(Task.created_at >= datetime(current_year, 1, 1)).count()

    stats = {
        'total': total_tasks,
        'completed': completed_tasks,
        'overdue': overdue_tasks,
        'in_progress': in_progress_tasks,
        'percentage': percentage,
        'current_year': current_year,
        'year_docs': year_docs,
        'year_tasks': year_tasks
    }

    # Документы и «сирот»-задачи
    if search_q:
        like_q = f'%{search_q}%'
        documents = FileDocument.query.filter(
            db.or_(
                FileDocument.filename.ilike(like_q),
                FileDocument.doc_number.ilike(like_q),
            )
        ).order_by(FileDocument.uploaded_at.desc()).all()
        orphan_tasks = Task.query.filter(
            Task.document_id.is_(None),
            db.or_(
                Task.title.ilike(like_q),
                Task.text.ilike(like_q),
            )
        ).all()
    else:
        documents = FileDocument.query.order_by(FileDocument.uploaded_at.desc()).all()
        orphan_tasks = Task.query.filter_by(document_id=None).all()

    settings = NotificationSettings.query.first()
    deadline_colors_enabled = True if not settings else settings.enable_deadline_colors

    district_names = _get_district_names()

    # Enhance documents with their task summaries
    for doc in documents:
        doc.total = len(doc.tasks)
        doc.completed = sum(1 for t in doc.tasks if t.status == 'Выполнено')
        doc.overdue = sum(1 for t in doc.tasks if t.status == 'Просрочено')
        for t in doc.tasks:
            t.deadline_flag = _deadline_flag(t, deadline_colors_enabled)

        # В дашборде не показываем все районы поштучно:
        # для массовых районных поручений — одна строка, счётчик уменьшается по мере сдачи отчётов.
        mass_buckets = defaultdict(list)
        doc_tasks_view = []
        for t in doc.tasks:
            is_district_task = t.executor and t.executor.name in district_names
            if not is_district_task:
                setattr(t, 'group_executor', t.executor.name if t.executor else 'Не назначен')
                setattr(t, 'mass_view', False)
                setattr(t, 'mass_count', 1)
                setattr(t, 'display_status', t.status)
                doc_tasks_view.append(t)
                continue
            mass_buckets[_mass_group_key(t)].append(t)

        for group_tasks in mass_buckets.values():
            doc_tasks_view.append(_mass_task_view_stats(group_tasks))

        setattr(doc, 'tasks_view', doc_tasks_view)

    active_documents = [d for d in documents if d.total == 0 or d.completed < d.total]
    completed_documents = [d for d in documents if d.total > 0 and d.completed >= d.total]

    for t in orphan_tasks:
        t.deadline_flag = _deadline_flag(t, deadline_colors_enabled)

    return render_template('dashboard.html',
                           tasks=tasks,
                           stats=stats,
                           documents=active_documents,
                           completed_documents=completed_documents,
                           orphan_tasks=orphan_tasks,
                           search_q=search_q,
                           tomorrow=datetime.now().date(),
                           deadline_colors_enabled=deadline_colors_enabled,
                           debug_mode=debug_mode)

@main.route('/leaderboard')
def leaderboard():
    leaders, traffic_stats = _build_leaderboard_data()
    return render_template(
        'leaderboard.html',
        leaders=leaders,
        traffic_stats=traffic_stats,
        district_links=_get_district_links(),
    )

@main.route('/kchz-traffic-light')
def traffic_light():
    return render_template('kchz_traffic_light.html')


@main.route('/admin/settings')
@require_admin('tasks')
def admin_settings():
    settings = NotificationSettings.query.first()
    if not settings:
        settings = NotificationSettings()
        db.session.add(settings)
        db.session.commit()
    recipients = NotificationRecipient.query.order_by(NotificationRecipient.created_at.desc()).all()
    return render_template('admin_settings.html', settings=settings, recipients=recipients)

@main.route('/add', methods=['GET'])
def add_task_page():
    executors = Executor.query.order_by(Executor.name).all()
    return render_template('add_task.html', executors=executors)

@main.route('/add', methods=['POST'])
@require_write('tasks')
def add_task():
    try:
        data = request.form
        executor_name = data.get('executor', '').strip()
        if not executor_name:
            return jsonify({"error": "Не указан исполнитель"}), 400

        executor = Executor.query.filter_by(name=executor_name).first()
        if not executor:
            executor = Executor(name=executor_name)
            db.session.add(executor)
            db.session.flush()

        deadline_str = data.get('deadline', '')
        try:
            deadline = datetime.strptime(deadline_str, '%Y-%m-%d')
        except ValueError:
            deadline = datetime.now() + timedelta(days=30)

        new_task = Task(
            item_number=data.get('item_number', 'Б/Н'),
            title=data.get('title', 'Без названия'),
            text=data.get('text', ''),
            deadline=deadline,
            file_hash=hashlib.md5(f"manual-{datetime.now().isoformat()}".encode()).hexdigest(),
            executor_id=executor.id,
            status='В работе'
        )
        db.session.add(new_task)
        db.session.commit()
        return jsonify({"ok": True, "message": "Поручение добавлено", "task_id": new_task.id}), 201
    except Exception as e:
        current_app.logger.error(f"Error adding task: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# ─── Загрузка PDF ─────────────────────────────────────────────────────

@main.route('/upload', methods=['POST'])
@require_write('tasks')
def upload_protocol():
    if 'document' in request.files:
        file = request.files['document']
    elif 'file' in request.files:
        file = request.files['file']
    else:
        return jsonify({"error": "Файл не найден"}), 400
    try:
        file_bytes = file.read()
        if not file_bytes:
            return jsonify({"error": "Пустой файл"}), 400

        file_hash = hashlib.md5(file_bytes).hexdigest()
    except Exception as e:
        current_app.logger.error(f"Error reading upload file: {e}")
        return jsonify({"error": "Ошибка чтения файла"}), 500

    # 1. Сохраняем документ для истории
    new_doc = FileDocument.query.filter_by(file_hash=file_hash).first()
    if not new_doc:
        new_doc = FileDocument(
            filename=file.filename,
            file_hash=file_hash,
            file_data=file_bytes
        )
        db.session.add(new_doc)
        db.session.flush()

    # 2. Запускаем наш парсер
    try:
        result = parse_pdf(file_bytes=file_bytes, filename=file.filename)
    except Exception as e:
        current_app.logger.error(f"Critical error in parse_pdf: {e}", exc_info=True)
        return jsonify({"error": f"Ошибка парсинга: {str(e)}"}), 500

    if not result or not result.get('tasks'):
        current_app.logger.warning(f"Parse result empty for file {file.filename}")
        return jsonify({"error": "Не удалось распарсить PDF или файлы не содержат задач"}), 500

    # Обновляем метаданные документа
    new_doc.doc_number = result.get('doc_number')
    new_doc.doc_date = result.get('doc_date')

    parsed_tasks = result['tasks']
    new_tasks_count = 0

    for t_data in parsed_tasks:
        # 1. Проверяем/Создаем исполнителя
        executor_name = t_data['executor']
        executor = Executor.query.filter_by(name=executor_name).first()
        if not executor:
            executor = Executor(name=executor_name)
            db.session.add(executor)
            db.session.flush() # Получаем ID

        # 2. Проверяем дубликаты (хэш файла + номер пункта + ID исполнителя)
        item_num = t_data.get('item_number', t_data.get('title'))
        exists = Task.query.filter_by(
            file_hash=file_hash,
            item_number=item_num,
            executor_id=executor.id
        ).first()

        if not exists:
            new_task = Task(
                item_number=item_num,
                title=t_data['title'],
                text=t_data.get('text', ''),
                deadline=t_data['deadline'],
                file_hash=file_hash,
                executor_id=executor.id,
                document_id=new_doc.id
            )
            db.session.add(new_task)
            new_tasks_count += 1

    db.session.commit()
    return jsonify({
        "message": f"Обработано. Добавлено {new_tasks_count} новых пунктов.",
        "doc_number": new_doc.doc_number,
        "new_count": new_tasks_count
    }), 201

# ─── API: Скачивание документа ────────────────────────────────────────

@main.route('/api/document/<int:doc_id>/download')
def download_document(doc_id):
    doc = FileDocument.query.get_or_404(doc_id)
    return send_file(
        io.BytesIO(doc.file_data),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=doc.filename or f'document_{doc_id}.pdf'
    )

# ─── API: Переключение отчёта ────────────────────────────────────────

@main.route('/<int:task_id>/report', methods=['POST'])
@require_write('tasks')
def toggle_report(task_id):
    task = Task.query.get_or_404(task_id)
    task.report_submitted = not task.report_submitted
    db.session.commit()
    return jsonify({"ok": True, "report_submitted": task.report_submitted})

# ─── API: Обновление статуса ─────────────────────────────────────────

@main.route('/<int:task_id>/status', methods=['POST'])
@require_write('tasks')
def update_status(task_id):
    task = Task.query.get_or_404(task_id)
    data = request.get_json(silent=True) or {}
    new_status = data.get('status', 'В работе')

    # Если помечаем как выполненную — требуем отчёт
    if new_status == 'Выполнено' and not task.report_submitted:
        return jsonify({"ok": False, "need_report": True, "error": "Сначала внесите отчёт"}), 400

    task.status = new_status
    db.session.commit()
    return jsonify({"ok": True, "status": task.status})

def _delete_document(doc):
    """Удалить документ и связанные поручения, не нарушая FK в PostgreSQL."""
    task_ids = [
        task_id for (task_id,) in db.session.query(Task.id).filter_by(document_id=doc.id).all()
    ]
    if task_ids:
        db.session.execute(
            delete(TaskNotificationLog).where(TaskNotificationLog.task_id.in_(task_ids))
        )
        db.session.flush()

        db.session.execute(
            update(YandexReport)
            .where(YandexReport.task_id.in_(task_ids))
            .values(task_id=None)
        )
        db.session.execute(
            delete(YandexReportTaskLink).where(YandexReportTaskLink.task_id.in_(task_ids))
        )
        db.session.flush()

        db.session.execute(
            delete(Task).where(Task.id.in_(task_ids))
        )
        db.session.flush()

    db.session.delete(doc)


def _delete_error_message(exc):
    message = str(exc)
    if 'ForeignKeyViolation' in message or 'foreign key constraint' in message.lower():
        return 'Не удалось удалить документ: есть связанные записи. Попробуйте ещё раз.'
    return message


# ─── API: Удаление документа (POST) ─────────────────────────────────

@main.route('/document/<int:doc_id>/delete', methods=['POST'])
@require_write('tasks')
def delete_document_post(doc_id):
    doc = FileDocument.query.get_or_404(doc_id)
    try:
        _delete_document(doc)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting doc {doc_id}: {e}")
        return jsonify({"ok": False, "error": _delete_error_message(e)}), 500

# ─── API: Удаление документа (DELETE) ────────────────────────────────

@main.route('/api/document/<int:doc_id>', methods=['DELETE'])
@require_write('tasks')
def delete_document_api(doc_id):
    doc = FileDocument.query.get_or_404(doc_id)
    try:
        _delete_document(doc)
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting doc {doc_id}: {e}")
        return jsonify({"ok": False, "error": _delete_error_message(e)}), 500

# ─── API: Экспорт Excel ──────────────────────────────────────────────

@main.route('/export/excel')
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Поручения"

        # Стили заголовков
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ['№', 'Документ', 'Пункт', 'Поручение', 'Исполнитель', 'Срок', 'Статус', 'Отчёт']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Данные
        tasks = Task.query.order_by(Task.deadline).all()
        for i, task in enumerate(tasks, 2):
            doc_name = task.document.filename if task.document else 'Ручной ввод'
            row_data = [
                i - 1,
                doc_name,
                task.item_number or '',
                task.title or '',
                task.executor.name if task.executor else '',
                task.deadline.strftime('%d.%m.%Y') if task.deadline else '',
                task.status or '',
                'Да' if task.report_submitted else 'Нет'
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border

        # Ширины столбцов
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 10

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Porycheniya.xlsx'
        )
    except Exception as e:
        current_app.logger.error(f"Error exporting Excel: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ─── API: Аналитика ──────────────────────────────────────────────────

@main.route('/api/analytics')
@require_read('tasks')
def analytics():
    period = request.args.get('period', 'month')
    period_value = request.args.get('period_value', '')

    now = datetime.now()

    # Определяем дату начала и конца периода
    if period == 'month' and '-' in period_value:
        parts = period_value.split('-')
        year = int(parts[0])
        month = int(parts[1])
        start_date = datetime(year, month + 1, 1)  # month is 0-indexed from JS
        if month + 1 == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 2, 1)
    elif period == 'quarter' and '-Q' in period_value:
        parts = period_value.split('-Q')
        year = int(parts[0])
        quarter = int(parts[1])
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)
        end_month = start_month + 3
        if end_month > 12:
            end_date = datetime(year + 1, end_month - 12, 1)
        else:
            end_date = datetime(year, end_month, 1)
    elif period == 'year':
        year = int(period_value) if period_value.isdigit() else now.year
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
    elif period == 'all':
        start_date = None
        end_date = None
    else:
        start_date = datetime(now.year, now.month, 1)
        if now.month == 12:
            end_date = datetime(now.year + 1, 1, 1)
        else:
            end_date = datetime(now.year, now.month + 1, 1)

    # Задачи за период
    if start_date is None:
        period_tasks = Task.query.all()
    else:
        period_tasks = Task.query.filter(
            Task.created_at >= start_date,
            Task.created_at < end_date
        ).all()

    total = len(period_tasks)
    completed = sum(1 for t in period_tasks if t.status == 'Выполнено')
    overdue = sum(1 for t in period_tasks if t.status == 'Просрочено')
    in_progress = total - completed - overdue
    pct = int(completed / total * 100) if total > 0 else 0

    # Месячная динамика (6 месяцев назад)
    monthly = []
    for i in range(5, -1, -1):
        m_date = datetime(now.year, now.month, 1) - timedelta(days=30 * i)
        m_start = datetime(m_date.year, m_date.month, 1)
        if m_date.month == 12:
            m_end = datetime(m_date.year + 1, 1, 1)
        else:
            m_end = datetime(m_date.year, m_date.month + 1, 1)

        m_created = Task.query.filter(Task.created_at >= m_start, Task.created_at < m_end).count()
        m_completed = Task.query.filter(
            Task.created_at >= m_start, Task.created_at < m_end,
            Task.status == 'Выполнено'
        ).count()

        months_ru = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
        monthly.append({
            'label': months_ru[m_start.month - 1] + ' ' + str(m_start.year),
            'created': m_created,
            'completed': m_completed
        })

    # Топ исполнителей
    top_executors = []
    executors = Executor.query.all()
    for ex in executors:
        ex_tasks = [t for t in period_tasks if t.executor_id == ex.id]
        ex_total = len(ex_tasks)
        if ex_total == 0:
            continue
        ex_completed = sum(1 for t in ex_tasks if t.status == 'Выполнено')
        top_executors.append({
            'name': ex.name,
            'total': ex_total,
            'completed': ex_completed,
            'percentage': int(ex_completed / ex_total * 100) if ex_total > 0 else 0
        })
    top_executors.sort(key=lambda x: x['percentage'], reverse=True)

    # КЧС аналитика
    kchz_data = _get_kchz_analytics()

    return jsonify({
        'summary': {
            'total': total,
            'completed': completed,
            'overdue': overdue,
            'in_progress': in_progress,
            'percentage': pct
        },
        'monthly': monthly,
        'top_executors': top_executors[:10],
        'kchz': kchz_data
    })


def _get_kchz_analytics():
    """Сбор аналитики КЧС по исполнителям (районам)."""
    try:
        executors = Executor.query.all()
        district_names = _get_district_names()
        districts = []
        total_all = 0
        green = 0
        yellow = 0
        red = 0

        for ex in executors:
            if district_names and ex.name not in district_names:
                continue
            ex_total = Task.query.filter_by(executor_id=ex.id).count()
            if ex_total == 0:
                continue
            ex_done = Task.query.filter_by(executor_id=ex.id, status='Выполнено').count()
            pct = int(ex_done / ex_total * 100) if ex_total > 0 else 0

            districts.append({
                'name': ex.name,
                'total': ex_total,
                'completed': ex_done,
                'percentage': pct
            })

            total_all += 1
            if pct >= 90:
                green += 1
            elif pct >= 70:
                yellow += 1
            else:
                red += 1

        districts.sort(key=lambda x: x['percentage'], reverse=True)

        return {
            'districts': districts,
            'stats': {
                'total': total_all,
                'green': green,
                'yellow': yellow,
                'red': red
            }
        }
    except Exception as e:
        current_app.logger.error(f"Error in KChZ analytics: {e}")
        return {'districts': [], 'stats': {'total': 0, 'green': 0, 'yellow': 0, 'red': 0}}


@main.route('/yandex/sync', methods=['POST'])
@require_write('tasks')
def yandex_sync():
    try:
        from services.yandex_disk import scan_reports
        result = scan_reports(current_app._get_current_object())
        return jsonify(result), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Yandex sync error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _parse_yandex_report_items(raw_value):
    if not raw_value:
        return []
    try:
        return json.loads(raw_value)
    except (json.JSONDecodeError, TypeError):
        return []


def _parse_yandex_match_details(raw_value):
    if not raw_value:
        return {}
    try:
        return json.loads(raw_value)
    except (json.JSONDecodeError, TypeError):
        return {}


@main.route('/yandex/test', methods=['POST'])
@require_write('tasks')
def yandex_test():
    try:
        from services.yandex_disk import YandexDiskClient, YANDEX_TOKEN
        if not YANDEX_TOKEN:
            return jsonify({
                "ok": True,
                "user": "Без токена (режим публичных папок)",
                "total_space": 0,
                "used_space": 0
            })
        return jsonify(YandexDiskClient().test_connection()), 200
    except Exception as e:
        current_app.logger.error(f"Yandex test error: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@main.route('/api/yandex/districts-status')
def yandex_districts_status():
    try:
        from services.yandex_disk import scan_districts_status
        return jsonify(scan_districts_status())
    except Exception as e:
        current_app.logger.error(f"Districts status error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/yandex/reports')
def yandex_reports():
    from models import YandexReportTaskLink
    from services.yandex_disk import _report_status_label
    include_archived = request.args.get('include_archived', '').lower() in ('1', 'true', 'yes')
    reports = YandexReport.query.order_by(YandexReport.received_at.desc()).all()
    payload = []
    for r in reports:
        is_superseded = bool(r.superseded_by_id)
        is_complete = (r.completeness_status or 'none') == 'full'
        if is_superseded and not include_archived:
            continue
        parsed_items = _parse_yandex_report_items(r.parsed_item_numbers)
        details = _parse_yandex_match_details(r.match_details)
        status, status_label = _report_status_label(r.items_matched, parsed_items)
        links = YandexReportTaskLink.query.filter_by(report_id=r.id).all()
        confidences = [lnk.confidence for lnk in links if lnk.confidence is not None]
        avg_confidence = round(sum(confidences) / len(confidences), 2) if confidences else None
        payload.append({
            "id": r.id,
            "filename": r.filename,
            "sender": r.sender_name,
            "received_at": r.received_at.strftime('%d.%m.%Y %H:%M') if r.received_at else None,
            "executor": r.executor.name if r.executor else None,
            "task_id": r.task_id,
            "kchs_number": r.kchs_number,
            "parsed_items": parsed_items,
            "items_matched": r.items_matched or 0,
            "status": status,
            "status_label": status_label,
            "completeness_status": r.completeness_status or "none",
            "file_version": r.file_version or 1,
            "suggest_count": len(details.get("suggest") or []),
            "missing_count": len(details.get("missing_tasks") or []),
            "avg_confidence": avg_confidence,
            "is_complete": is_complete,
            "is_superseded": is_superseded,
            "yandex_path": r.yandex_path,
            "synced_at": r.synced_at.strftime('%d.%m.%Y %H:%M') if r.synced_at else None,
        })
    return jsonify(payload)


@main.route('/api/yandex/reports/<int:report_id>/preview')
def yandex_report_preview(report_id):
    try:
        from services.report_matcher import get_report_preview
        return jsonify(get_report_preview(report_id))
    except Exception as e:
        current_app.logger.error(f"Report preview error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/yandex/completeness')
def yandex_completeness():
    try:
        from services.report_matcher import compute_district_completeness
        return jsonify(compute_district_completeness())
    except Exception as e:
        current_app.logger.error(f"Completeness error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/admin/yandex/review-queue')
@require_admin('tasks')
def yandex_review_queue():
    try:
        from services.report_matcher import get_review_queue
        return jsonify(get_review_queue())
    except Exception as e:
        current_app.logger.error(f"Review queue error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/admin/yandex/auto-link', methods=['POST'])
@require_admin('tasks')
def yandex_auto_link():
    try:
        from services.report_matcher import auto_link_pending_reports, confirm_report_suggestions
        data = request.get_json(silent=True) or {}
        report_id = data.get("report_id")
        if report_id:
            result, code = confirm_report_suggestions(
                report_id, include_suggestions=data.get("include_suggestions", True)
            )
            return jsonify(result), code
        result = auto_link_pending_reports(
            current_app._get_current_object(),
            include_suggestions=data.get("include_suggestions", True),
        )
        return jsonify({"ok": True, **result})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Auto-link error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/admin/yandex/link', methods=['POST'])
@require_admin('tasks')
def yandex_manual_link():
    try:
        from services.report_matcher import manual_link_report
        data = request.get_json(silent=True) or {}
        report_id = data.get("report_id")
        task_ids = data.get("task_ids") or []
        if not report_id or not task_ids:
            return jsonify({"error": "Нужны report_id и task_ids"}), 400
        result, code = manual_link_report(report_id, task_ids)
        return jsonify(result), code
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Manual link error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/kchz-traffic-light')
def kchz_traffic_light():
    districts = []
    leaders, traffic_stats = _build_leaderboard_data()
    for idx, leader in enumerate(leaders, start=1):
        districts.append({
            "num": idx,
            "name": leader["name"],
            "total": leader["total"],
            "completed": leader["completed"],
            "percentage": leader["percentage"],
            "color": leader["color"],
            "in_time": leader["timing"]["done_in_time"],
            "late": leader["timing"]["done_late"],
            "no_deadline": leader["timing"]["done_no_deadline"],
        })
    return jsonify({"districts": districts, "stats": traffic_stats})


@main.route('/upload-excel-traffic', methods=['POST'])
@require_write('tasks')
def upload_excel_traffic():
    if 'file' not in request.files:
        return jsonify({"error": "Файл не найден"}), 400
    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({"error": "Требуется файл .xlsx"}), 400
    file.save(_excel_storage_path())
    return jsonify({"ok": True, "message": "Excel загружен"}), 201


@main.route('/download-excel-traffic')
def download_excel_traffic():
    storage_path = _excel_storage_path()
    if not os.path.exists(storage_path):
        return jsonify({"error": "Файл светофора не загружен"}), 404
    return send_file(storage_path, as_attachment=True, download_name=KCHZ_EXCEL_FILENAME)


@main.route('/delete-excel-traffic', methods=['DELETE'])
@require_write('tasks')
def delete_excel_traffic():
    storage_path = _excel_storage_path()
    if os.path.exists(storage_path):
        os.remove(storage_path)
    return jsonify({"ok": True})


@main.route('/api/admin/settings', methods=['GET', 'POST'])
@require_admin('tasks')
def api_admin_settings():
    settings = NotificationSettings.query.first()
    if not settings:
        settings = NotificationSettings()
        db.session.add(settings)
        db.session.flush()

    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        settings.enable_email = bool(data.get('enable_email', settings.enable_email))
        settings.enable_telegram = bool(data.get('enable_telegram', settings.enable_telegram))
        settings.enable_deadline_colors = bool(data.get('enable_deadline_colors', settings.enable_deadline_colors))
        settings.notify_days_7 = bool(data.get('notify_days_7', settings.notify_days_7))
        settings.notify_days_3 = bool(data.get('notify_days_3', settings.notify_days_3))
        db.session.commit()

    return jsonify({
        "enable_email": settings.enable_email,
        "enable_telegram": settings.enable_telegram,
        "enable_deadline_colors": settings.enable_deadline_colors,
        "notify_days_7": settings.notify_days_7,
        "notify_days_3": settings.notify_days_3,
    })


@main.route('/api/admin/recipients', methods=['GET', 'POST', 'DELETE'])
@require_admin('tasks')
def api_admin_recipients():
    if request.method == 'GET':
        recipients = NotificationRecipient.query.order_by(NotificationRecipient.created_at.desc()).all()
        return jsonify([{"id": r.id, "email": r.email, "is_active": r.is_active} for r in recipients])

    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        email = (data.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return jsonify({"error": "Некорректный email"}), 400
        existing = NotificationRecipient.query.filter_by(email=email).first()
        if existing:
            existing.is_active = True
            db.session.commit()
            return jsonify({"ok": True, "id": existing.id})
        recipient = NotificationRecipient(email=email, is_active=True)
        db.session.add(recipient)
        db.session.commit()
        return jsonify({"ok": True, "id": recipient.id}), 201

    data = request.get_json(silent=True) or {}
    rec_id = data.get("id")
    recipient = NotificationRecipient.query.get_or_404(rec_id)
    db.session.delete(recipient)
    db.session.commit()
    return jsonify({"ok": True})


KCHS_CLEAR_CONFIRM_PHRASE = "ОЧИСТИТЬ БАЗУ"


@main.route('/api/admin/clear-kchs', methods=['POST'])
@require_admin('tasks')
def api_admin_clear_kchs():
    data = request.get_json(silent=True) or {}
    confirm_phrase = (data.get("confirm_phrase") or "").strip().upper()
    if confirm_phrase != KCHS_CLEAR_CONFIRM_PHRASE:
        return jsonify({
            "error": f"Для подтверждения введите фразу: {KCHS_CLEAR_CONFIRM_PHRASE}",
        }), 400

    try:
        from services.kchs_maintenance import clear_kchs_database
        result = clear_kchs_database()
        return jsonify({"ok": True, **result})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Clear KCHS DB error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@main.route('/api/admin/smtp-status')
@require_admin('tasks')
def api_smtp_status():
    from services.notifier import get_smtp_status
    return jsonify(get_smtp_status())


@main.route('/api/admin/test-email', methods=['POST'])
@require_admin('tasks')
def api_test_email():
    from services.notifier import send_test_email
    data = request.get_json(silent=True) or {}
    result, code = send_test_email(data.get("email"))
    return jsonify(result), code


@main.route('/api/admin/run-notifications', methods=['POST'])
@require_admin('tasks')
def api_run_notifications():
    try:
        from services.notifier import check_deadlines_and_notify
        check_deadlines_and_notify(current_app._get_current_object())
        return jsonify({"ok": True})
    except Exception as e:
        current_app.logger.error(f"Notification run failed: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)}), 500


@main.route('/api/analytics/favorites')
def favorite_analytics():
    executor_name = request.args.get("executor", "").strip()
    period = request.args.get("period", "1m").strip().lower()
    if period not in {"1m", "3m", "6m", "1y"}:
        period = "1m"

    district_names = _get_district_names()
    favorites = Executor.query.order_by(Executor.name.asc()).all()
    favorites = [e for e in favorites if (not district_names or e.name not in district_names)]

    if not favorites:
        return jsonify({"favorites": [], "metrics": None})

    executor = None
    if executor_name:
        executor = next((e for e in favorites if e.name == executor_name), None)
    if not executor:
        executor = next((e for e in favorites if "лепчиков" in e.name.lower()), favorites[0])

    metrics, _ = _favorite_metrics(executor, period)
    return jsonify({
        "favorites": [e.name for e in favorites],
        "metrics": metrics
    })


@main.route('/api/analytics/favorites/export')
def favorite_analytics_export():
    executor_name = request.args.get("executor", "").strip()
    period = request.args.get("period", "1m").strip().lower()
    if not executor_name:
        return jsonify({"error": "Не указан исполнитель"}), 400

    executor = Executor.query.filter_by(name=executor_name).first()
    if not executor:
        return jsonify({"error": "Исполнитель не найден"}), 404

    metrics, tasks = _favorite_metrics(executor, period if period in {"1m", "3m", "6m", "1y"} else "1m")

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Избранное"

    ws["A1"] = f"Аналитика: {metrics['executor']}"
    ws["A2"] = f"Период: {metrics['from']} - {metrics['to']}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True)

    rows = [
        ("Всего поручений", metrics["total"]),
        ("Выполнено", metrics["completed"]),
        ("В срок", metrics["in_time"]),
        ("С опозданием", metrics["late"]),
        ("Без срока", metrics["no_deadline"]),
        ("Открыто", metrics["open"]),
        ("Рейтинг (в срок)", f"{metrics['performance_pct']}%"),
        ("Светофор", metrics["traffic_color"]),
    ]
    for idx, (label, value) in enumerate(rows, start=4):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value

    header_row = 14
    ws[f"A{header_row}"] = "ID"
    ws[f"B{header_row}"] = "Пункт"
    ws[f"C{header_row}"] = "Поручение"
    ws[f"D{header_row}"] = "Срок"
    ws[f"E{header_row}"] = "Статус"
    ws[f"F{header_row}"] = "Оценка"
    for col in "ABCDEF":
        ws[f"{col}{header_row}"].font = Font(bold=True)

    for i, task in enumerate(tasks, start=header_row + 1):
        ws[f"A{i}"] = task.id
        ws[f"B{i}"] = task.item_number or ""
        ws[f"C{i}"] = task.title or ""
        deadline = _safe_deadline(task)
        ws[f"D{i}"] = deadline.strftime("%d.%m.%Y") if deadline else "Нет срока"
        ws[f"E{i}"] = task.status
        ws[f"F{i}"] = _task_timing_status(task)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"favorite_{executor.name}_{period}.xlsx"
    )