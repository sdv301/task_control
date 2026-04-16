import os
import sys
import logging
import time
import hashlib
import io
import platform
from datetime import datetime, timedelta
from collections import OrderedDict

from flask import Flask, render_template, jsonify, request, redirect, url_for, send_file
from models import db, Task, Executor, FileDocument, YandexReport
from pdf_parser.pdf_engine import parse_pdf
from routes import main as main_blueprint

# ─── Кроссплатформенные пути ──────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, '..', 'data')
DATA_DIR = os.path.abspath(DATA_DIR)

# ─── Логирование ─────────────────────────────────────────────────────────────
LOG_DIR = os.path.join(DATA_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(LOG_DIR, 'app.log'),
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

app = Flask(__name__)

# Определяем DATABASE_URL с учётом платформы
_default_db = os.environ.get('DATABASE_URL')
if not _default_db:
    if platform.system() == 'Windows':
        db_path = os.path.join(DATA_DIR, 'tasks.db').replace('\\', '/')
        _default_db = f'sqlite:///{db_path}'
    else:
        _default_db = 'sqlite:////data/tasks.db'

app.config['SQLALCHEMY_DATABASE_URI'] = _default_db
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-prod')
db.init_app(app)

# ─── Регистрация Blueprint ─────────────────────────────────────────────────────
app.register_blueprint(main_blueprint)

# ─── Инициализация БД ────────────────────────────────────────────────────────
with app.app_context():
    os.makedirs(DATA_DIR, exist_ok=True)
    db.create_all()
    logging.info("БД инициализирована успешно.")


def _refresh_statuses(tasks, now):
    changed = False
    for task in tasks:
        if task.status != "Выполнено" and task.deadline < now:
            task.status = "Просрочено"
            changed = True
    if changed:
        db.session.commit()


# ─── Главная страница ─────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


# ─── Загрузка PDF ─────────────────────────────────────────────────────────────
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'document' not in request.files:
        return jsonify({"error": "Файл не найден в запросе"}), 400

    file = request.files['document']
    if not file or file.filename == '':
        return jsonify({"error": "Файл не выбран"}), 400

    file_bytes = file.read()
    file_hash = hashlib.md5(file_bytes).hexdigest()

    # Сохранение документа
    new_doc = FileDocument.query.filter_by(file_hash=file_hash).first()
    if not new_doc:
        new_doc = FileDocument(filename=file.filename, file_data=file_bytes, file_hash=file_hash)
        db.session.add(new_doc)
        db.session.flush()

    # Парсинг
    result = parse_pdf(file_bytes=file_bytes, filename=file.filename)
    if not result:
        db.session.rollback()
        return jsonify({"error": "Не удалось распознать текст из PDF. Возможно, это скан-копия, а проект запущен локально без OCR (Tesseract). Пожалуйста, запустите проект через Docker."}), 422

    # Сохраняем номер и дату документа
    parsed_tasks = result.get('tasks', [])
    new_doc.doc_number = result.get('doc_number')
    new_doc.doc_date = result.get('doc_date')

    new_count = 0
    for t_data in parsed_tasks:
        executor_name = t_data['executor']
        executor = Executor.query.filter_by(name=executor_name).first()
        if not executor:
            executor = Executor(name=executor_name)
            db.session.add(executor)
            db.session.flush()

        item_num = t_data.get('item_number', t_data['title'])
        exists = Task.query.filter_by(
            file_hash=file_hash,
            item_number=item_num,
            executor_id=executor.id
        ).first()

        if not exists:
            new_task = Task(
                item_number=item_num,
                title=t_data['title'],
                text=t_data.get('text', ''),
                deadline=t_data['deadline'],
                file_hash=file_hash,
                executor_id=executor.id,
                document_id=new_doc.id
            )
            db.session.add(new_task)
            new_count += 1

    db.session.commit()
    logging.info(f"Загружен файл '{file.filename}': добавлено {new_count} новых задач.")
    return jsonify({"message": f"Успешно! Добавлено {new_count} новых поручений.", "count": new_count}), 201


# ─── Дашборд ──────────────────────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    search_q = request.args.get('q', '').strip()

    all_tasks = Task.query.order_by(Task.deadline.asc()).all()
    now = datetime.utcnow()
    tomorrow = now + timedelta(days=1)
    _refresh_statuses(all_tasks, now)

    # Применяем поиск
    if search_q:
        q_lower = search_q.lower()
        filtered_tasks = [
            t for t in all_tasks
            if q_lower in (t.title or '').lower()
            or q_lower in (t.text or '').lower()
            or q_lower in (t.executor.name or '').lower()
            or (t.document and q_lower in (t.document.filename or '').lower())
        ]
    else:
        filtered_tasks = all_tasks

    # Группировка по документам
    documents = OrderedDict()
    orphan_tasks = []

    try:
        from services.districts import DISTRICTS
        district_names = set(DISTRICTS.keys())
    except ImportError:
        district_names = set()

    for task in filtered_tasks:
        if task.document_id and task.document:
            doc_id = task.document_id
            if doc_id not in documents:
                doc = task.document
                documents[doc_id] = {
                    'id': doc.id,
                    'filename': doc.filename,
                    'uploaded_at': doc.uploaded_at,
                    'doc_number': doc.doc_number,
                    'doc_date': doc.doc_date,
                    'tasks': [],
                    'title_counts': {},
                    'total': 0,
                    'overdue': 0,
                    'completed': 0,
                }
            documents[doc_id]['tasks'].append(task)
            documents[doc_id]['total'] += 1
            
            # Подсчет дублирующихся заголовков
            if task.executor and task.executor.name in district_names:
                t = task.title
                documents[doc_id]['title_counts'][t] = documents[doc_id]['title_counts'].get(t, 0) + 1
            
            if task.status == 'Просрочено':
                documents[doc_id]['overdue'] += 1
            elif task.status == 'Выполнено':
                documents[doc_id]['completed'] += 1
        else:
            orphan_tasks.append(task)

    # Проставляем визуальные свойства (setattr безопасен — используется только для рендеринга шаблона, не сохраняется в БД)
    for doc_id, doc_data in documents.items():
        doc_data['grouped_tasks'] = {}
        for task in doc_data['tasks']:
            is_mass = False
            if task.executor and task.executor.name in district_names:
                if doc_data['title_counts'].get(task.title, 0) > 10:
                    is_mass = True
            
            task.__dict__['is_mass_task'] = is_mass
            task.__dict__['display_executor'] = "Органам местного самоуправления Республики Саха (Якутия)" if is_mass else (task.executor.name if task.executor else "Не назначен")

    doc_list = list(documents.values())

    # Статистика по ВСЕМ (не отфильтрованным)
    total = len(all_tasks)
    overdue = sum(1 for t in all_tasks if t.status == 'Просрочено')
    completed = sum(1 for t in all_tasks if t.status == 'Выполнено')
    in_progress = total - overdue - completed
    percentage = int((completed / total * 100) if total > 0 else 0)

    # Статистика за год
    current_year = now.year
    year_tasks = [t for t in all_tasks if t.created_at and t.created_at.year == current_year]
    # Кроссплатформенный подсчёт документов за год (без db.extract для совместимости с SQLite)
    all_docs = FileDocument.query.all()
    year_docs = sum(1 for d in all_docs if d.uploaded_at and d.uploaded_at.year == current_year)

    stats = {
        'total': total,
        'overdue': overdue,
        'completed': completed,
        'percentage': percentage,
        'in_progress': in_progress,
        'year_docs': year_docs,
        'year_tasks': len(year_tasks),
        'current_year': current_year,
    }

    return render_template(
        'dashboard.html',
        documents=doc_list,
        orphan_tasks=orphan_tasks,
        stats=stats,
        search_q=search_q,
        tomorrow=tomorrow.date(),
        today=now.date()
    )


# ─── Добавить задачу вручную ──────────────────────────────────────────────────
@app.route('/task/add', methods=['GET', 'POST'])
def add_task():
    if request.method == 'POST':
        executor_name = request.form.get('executor', '').strip()
        title        = request.form.get('title', '').strip()
        text         = request.form.get('text', '').strip()
        deadline_str = request.form.get('deadline', '').strip()

        if not executor_name or not title or not deadline_str:
            return render_template('add_task.html', error="Заполните все обязательные поля.")

        try:
            deadline = datetime.strptime(deadline_str, '%Y-%m-%d')
        except ValueError:
            return render_template('add_task.html', error="Неверный формат даты.")

        executor = Executor.query.filter_by(name=executor_name).first()
        if not executor:
            executor = Executor(name=executor_name)
            db.session.add(executor)
            db.session.flush()

        fake_hash = hashlib.sha256(f"manual-{title}-{datetime.utcnow().isoformat()}".encode()).hexdigest()
        task = Task(
            item_number="Ручной ввод",
            title=title,
            text=text,
            deadline=deadline,
            file_hash=fake_hash,
            executor_id=executor.id
        )
        db.session.add(task)
        db.session.commit()
        logging.info(f"Задача '{title}' добавлена вручную.")
        return redirect(url_for('dashboard'))

    return render_template('add_task.html')


# ─── Изменить статус задачи ───────────────────────────────────────────────────
@app.route('/task/<int:task_id>/status', methods=['POST'])
def update_status(task_id):
    task = Task.query.get_or_404(task_id)
    new_status = request.json.get('status')
    if new_status == 'Выполнено' and not task.report_submitted:
        return jsonify({"error": "Сначала внесите отчёт", "need_report": True}), 400
    if new_status in ('В работе', 'Выполнено', 'Просрочено'):
        task.status = new_status
        db.session.commit()
        return jsonify({"ok": True, "status": new_status})
    return jsonify({"error": "Недопустимый статус"}), 400


@app.route('/task/<int:task_id>/report', methods=['POST'])
def toggle_report(task_id):
    task = Task.query.get_or_404(task_id)
    task.report_submitted = not task.report_submitted
    db.session.commit()
    return jsonify({"ok": True, "report_submitted": task.report_submitted})


@app.route('/document/<int:doc_id>/delete', methods=['POST'])
def delete_document(doc_id):
    doc = FileDocument.query.get_or_404(doc_id)
    # Удаляем все задачи этого документа
    Task.query.filter_by(document_id=doc_id).delete()
    db.session.delete(doc)
    db.session.commit()
    logging.info(f"Удалён документ '{doc.filename}' (id={doc_id}) и все его задачи.")
    return jsonify({"ok": True})


# ─── Экспорт в Excel ──────────────────────────────────────────────────────────
@app.route('/export/excel')
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "Модуль openpyxl не установлен."}), 500

    tasks = Task.query.order_by(Task.deadline.asc()).all()
    now = datetime.utcnow()
    _refresh_statuses(tasks, now)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Поручения"

    CLR_HEADER_BG  = "1E3A5F"
    CLR_HEADER_FG  = "FFFFFF"
    CLR_ROW_ODD    = "EBF1F8"
    CLR_ROW_EVEN   = "FFFFFF"
    CLR_OVERDUE    = "FFD7D7"
    CLR_COMPLETED  = "D4EDDA"
    CLR_TITLE_BG   = "2E75B6"
    CLR_SUBTTL_BG  = "D6E4F0"

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "ИСПОЛНЕНИЕ ПОРУЧЕНИЙ"
    title_cell.font = Font(bold=True, size=14, color=CLR_HEADER_FG, name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=CLR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    date_cell = ws['A2']
    date_cell.value = f"Сформировано: {now.strftime('%d.%m.%Y %H:%M')} UTC"
    date_cell.font = Font(italic=True, size=10, color="555555", name="Calibri")
    date_cell.fill = PatternFill("solid", fgColor=CLR_SUBTTL_BG)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HEADERS = ["No", "Поручение / Пункт", "Текст поручения", "Исполнитель", "Срок исполнения", "Статус", "Дней до срока"]
    WIDTHS  = [5, 22, 50, 24, 18, 14, 14]

    for col_idx, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11, color=CLR_HEADER_FG, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"

    for row_idx, task in enumerate(tasks, start=1):
        excel_row = row_idx + 4
        deadline_date = task.deadline.date()
        days_left = (deadline_date - now.date()).days

        if task.status == 'Просрочено':
            row_color = CLR_OVERDUE
        elif task.status == 'Выполнено':
            row_color = CLR_COMPLETED
        else:
            row_color = CLR_ROW_ODD if row_idx % 2 == 1 else CLR_ROW_EVEN

        fill = PatternFill("solid", fgColor=row_color)
        days_str = f"+{days_left}" if days_left >= 0 else str(days_left)

        row_data = [
            row_idx,
            task.title or "-",
            task.text or "-",
            task.executor.name if task.executor else "-",
            deadline_date.strftime('%d.%m.%Y'),
            task.status,
            days_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border
            cell.font = Font(size=10, name="Calibri")
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == 3),
                horizontal="center" if col_idx not in (2, 3, 4) else "left"
            )
            if col_idx == 7 and days_left < 0:
                cell.font = Font(size=10, name="Calibri", color="CC0000", bold=True)

        ws.row_dimensions[excel_row].height = 30

    total = len(tasks)
    completed_cnt = sum(1 for t in tasks if t.status == 'Выполнено')
    overdue_cnt   = sum(1 for t in tasks if t.status == 'Просрочено')
    inprog_cnt    = total - completed_cnt - overdue_cnt

    summary_row = len(tasks) + 5
    ws.merge_cells(f'A{summary_row}:C{summary_row}')
    ws[f'A{summary_row}'].value = f"ИТОГО: {total} поручений"
    ws[f'A{summary_row}'].font = Font(bold=True, size=11, name="Calibri")
    ws[f'A{summary_row}'].fill = PatternFill("solid", fgColor=CLR_SUBTTL_BG)
    ws[f'A{summary_row}'].alignment = Alignment(horizontal="left", vertical="center")
    ws[f'A{summary_row}'].border = border

    ws[f'D{summary_row}'].value = f"В работе: {inprog_cnt}"
    ws[f'E{summary_row}'].value = f"Выполнено: {completed_cnt}"
    ws[f'F{summary_row}'].value = f"Просрочено: {overdue_cnt}"
    for col in ['D', 'E', 'F', 'G']:
        ws[f'{col}{summary_row}'].font = Font(bold=True, size=10, name="Calibri")
        ws[f'{col}{summary_row}'].fill = PatternFill("solid", fgColor=CLR_SUBTTL_BG)
        ws[f'{col}{summary_row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'{col}{summary_row}'].border = border
    ws.row_dimensions[summary_row].height = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Porycheniya_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ─── Яндекс.Диск: синхронизация ──────────────────────────────────────────────
@app.route('/yandex/sync', methods=['POST'])
def yandex_sync():
    """Ручная синхронизация с Яндекс.Диском"""
    try:
        from services.yandex_disk import scan_reports
        results = scan_reports(app)
        logging.info(f"Синхронизация Я.Диска: {results['new']} новых, {results['skipped']} пропущено.")
        return jsonify(results), 200
    except Exception as e:
        logging.error(f"Ошибка синхронизации Я.Диска: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/yandex/test', methods=['POST'])
def yandex_test_connection():
    """Проверка подключения к Яндекс.Диску"""
    try:
        from services.yandex_disk import YandexDiskClient, YANDEX_TOKEN
        if not YANDEX_TOKEN:
            return jsonify({
                "ok": True, 
                "user": "Без токена (Режим публичных папок)",
                "total_space": 0,
                "used_space": 0
            })
        client = YandexDiskClient()
        result = client.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route('/api/yandex/reports')
def api_yandex_reports():
    """API: список отчётов с Яндекс.Диска"""
    reports = YandexReport.query.order_by(YandexReport.received_at.desc()).all()
    return jsonify([{
        "id": r.id,
        "filename": r.filename,
        "sender": r.sender_name,
        "received_at": r.received_at.strftime('%d.%m.%Y %H:%M') if r.received_at else None,
        "executor": r.executor.name if r.executor else None,
        "task_id": r.task_id,
        "synced_at": r.synced_at.strftime('%d.%m.%Y %H:%M') if r.synced_at else None,
    } for r in reports])


# ─── Страница «Светофор» ─────────────────────────────────────────────────────
@app.route('/leaderboard')
def leaderboard():
    """Рейтинг исполнителей с цветовой индикацией выполнения"""
    all_tasks = Task.query.all()
    now = datetime.utcnow()
    _refresh_statuses(all_tasks, now)

    executors = Executor.query.all()
    leaders = []

    # Подгружаем список районов для фильтрации
    try:
        from services.districts import DISTRICTS
        district_names = set(DISTRICTS.keys())
    except ImportError:
        district_names = set()

    for ex in executors:
        if district_names and ex.name not in district_names:
            continue

        tasks = [t for t in all_tasks if t.executor_id == ex.id]
        total = len(tasks)
        completed = sum(1 for t in tasks if t.status == 'Выполнено')
        overdue = sum(1 for t in tasks if t.status == 'Просрочено')
        in_progress = total - completed - overdue
        reports_count = YandexReport.query.filter_by(executor_id=ex.id).count()

        # Последний отчёт с Яндекс.Диска
        last_report = YandexReport.query.filter_by(executor_id=ex.id) \
            .order_by(YandexReport.received_at.desc()).first()

        # Логика светофора
        if total == 0:
            color = 'gray'
        elif completed == total:
            color = 'green'
        elif overdue > 0 and completed == 0:
            color = 'red'
        else:
            color = 'yellow'

        percentage = int(completed / total * 100) if total > 0 else 0

        leaders.append({
            'id': ex.id,
            'name': ex.name,
            'total': total,
            'completed': completed,
            'overdue': overdue,
            'in_progress': in_progress,
            'percentage': percentage,
            'color': color,
            'reports_count': reports_count,
            'last_report_date': last_report.received_at.strftime('%d.%m.%Y %H:%M') if last_report and last_report.received_at else None,
            'last_report_file': last_report.filename if last_report else None,
            'tasks': tasks,
        })

    # Сортировка: зелёные → жёлтые → красные → серые
    color_order = {'green': 0, 'yellow': 1, 'red': 2, 'gray': 3}
    leaders.sort(key=lambda x: (color_order.get(x['color'], 3), -x['percentage']))

    # Общая статистика для светофора
    total_executors = len(leaders)
    green_count = sum(1 for l in leaders if l['color'] == 'green')
    yellow_count = sum(1 for l in leaders if l['color'] == 'yellow')
    red_count = sum(1 for l in leaders if l['color'] == 'red')

    traffic_stats = {
        'total': total_executors,
        'green': green_count,
        'yellow': yellow_count,
        'red': red_count,
    }

    return render_template('leaderboard.html', leaders=leaders, traffic_stats=traffic_stats)


# ─── API: статистика ─────────────────────────────────────────────────────────
@app.route('/api/stats')
def api_stats():
    tasks = Task.query.all()
    now = datetime.utcnow()
    _refresh_statuses(tasks, now)
    total = len(tasks)
    completed = sum(1 for t in tasks if t.status == 'Выполнено')
    overdue   = sum(1 for t in tasks if t.status == 'Просрочено')
    return jsonify({
        "total": total,
        "completed": completed,
        "overdue": overdue,
        "in_progress": total - completed - overdue
    })


# ─── API: Расширенная аналитика ──────────────────────────────────────────────
@app.route('/api/analytics')
def api_analytics():
    """Расширенная аналитика с фильтрацией по периоду"""
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta

    period = request.args.get('period', 'month')
    period_value = request.args.get('period_value', '')

    now = datetime.utcnow()
    start_date = None
    end_date = now

    # Определяем период
    if period == 'month' and period_value:
        try:
            year, month = period_value.split('-')
            start_date = datetime(int(year), int(month), 1)
            if month == '12':
                end_date = datetime(int(year) + 1, 1, 1)
            else:
                end_date = datetime(int(year), int(month) + 1, 1)
        except (ValueError, TypeError):
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == 'quarter' and period_value:
        try:
            year, q = period_value.replace('Q', '-').split('-')
            q = int(q)
            start_date = datetime(int(year), (q - 1) * 3 + 1, 1)
            if q == 4:
                end_date = datetime(int(year) + 1, 1, 1)
            else:
                end_date = datetime(int(year), q * 3 + 1, 1)
        except (ValueError, TypeError):
            cur_q = (now.month - 1) // 3 + 1
            start_date = datetime(now.year, (cur_q - 1) * 3 + 1, 1)
    elif period == 'year' and period_value:
        try:
            start_date = datetime(int(period_value), 1, 1)
            end_date = datetime(int(period_value) + 1, 1, 1)
        except (ValueError, TypeError):
            start_date = datetime(now.year, 1, 1)
            end_date = datetime(now.year + 1, 1, 1)
    else:
        # По умолчанию — текущий месяц
        start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Все задачи
    all_tasks = Task.query.all()
    _refresh_statuses(all_tasks, now)

    # Фильтруем по периоду (по created_at)
    period_tasks = [t for t in all_tasks if t.created_at and start_date <= t.created_at < end_date]

    # Статистика за период
    total = len(period_tasks)
    completed = sum(1 for t in period_tasks if t.status == 'Выполнено')
    overdue = sum(1 for t in period_tasks if t.status == 'Просрочено')
    in_progress = total - completed - overdue
    pct = int((completed / total * 100) if total > 0 else 0)

    # Динамика по месяцам (последние 6 месяцев)
    monthly_data = []
    for i in range(5, -1, -1):
        d = datetime(now.year, now.month - i, 1) if now.month > i else datetime(now.year - 1, now.month + 12 - i, 1)
        if d.month == 12:
            next_d = datetime(d.year + 1, 1, 1)
        else:
            next_d = datetime(d.year, d.month + 1, 1)

        month_tasks = [t for t in all_tasks if t.created_at and d <= t.created_at < next_d]
        month_completed = sum(1 for t in month_tasks if t.status == 'Выполнено')
        month_names = ['Янв', 'Фев', 'Мар', 'Апр', 'Май', 'Июн', 'Июл', 'Авг', 'Сен', 'Окт', 'Ноя', 'Дек']
        monthly_data.append({
            'label': month_names[d.month - 1] + ' ' + str(d.year),
            'created': len(month_tasks),
            'completed': month_completed
        })

    # Топ исполнителей (реальные данные из БД)
    executors_data = []
    try:
        from services.districts import DISTRICTS
        district_names = set(DISTRICTS.keys())
    except ImportError:
        district_names = set()

    all_executors = Executor.query.all()
    for ex in all_executors:
        if district_names and ex.name not in district_names:
            continue
        ex_tasks = [t for t in all_tasks if t.executor_id == ex.id]
        ex_total = len(ex_tasks)
        if ex_total == 0:
            continue
        ex_completed = sum(1 for t in ex_tasks if t.status == 'Выполнено')
        ex_overdue = sum(1 for t in ex_tasks if t.status == 'Просрочено')
        ex_pct = int((ex_completed / ex_total * 100))
        executors_data.append({
            'name': ex.name,
            'total': ex_total,
            'completed': ex_completed,
            'overdue': ex_overdue,
            'percentage': ex_pct
        })

    # Сортировка по проценту выполнения
    executors_data.sort(key=lambda x: -x['percentage'])
    top_executors = executors_data[:10]

    # Статистика по КЧС (светофор)
    kchz_data = []
    try:
        import openpyxl
        xlsx_path = os.path.join(BASE_DIR, '..', 'светофоры.xlsx')
        xlsx_path = os.path.abspath(xlsx_path)
        if not os.path.exists(xlsx_path):
            xlsx_path = os.path.join(os.getcwd(), 'светофоры.xlsx')
        if not os.path.exists(xlsx_path):
            xlsx_path = os.path.join(BASE_DIR, 'светофоры.xlsx')

        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            for row in range(6, ws.max_row + 1):
                num = ws.cell(row=row, column=1).value
                name = ws.cell(row=row, column=2).value
                total_val = ws.cell(row=row, column=3).value
                completed_val = ws.cell(row=row, column=4).value
                percentage_val = ws.cell(row=row, column=5).value
                if not isinstance(num, (int, float)) or not name:
                    continue
                kchz_data.append({
                    'num': int(num),
                    'name': str(name).strip(),
                    'total': int(total_val) if total_val else 0,
                    'completed': int(completed_val) if completed_val else 0,
                    'percentage': round(float(percentage_val) * 100, 1) if percentage_val else 0,
                })
            wb.close()
    except Exception as e:
        logging.error(f"Ошибка чтения светофоры.xlsx для аналитики: {e}")

    # Сопоставление КЧС с задачами из БД
    for k in kchz_data:
        db_tasks = [t for t in all_tasks if t.executor and k['name'].lower() in t.executor.name.lower()]
        k['db_total'] = len(db_tasks)
        k['db_completed'] = sum(1 for t in db_tasks if t.status == 'Выполнено')
        k['db_overdue'] = sum(1 for t in db_tasks if t.status == 'Просрочено')
        k['db_percentage'] = int((k['db_completed'] / k['db_total'] * 100)) if k['db_total'] > 0 else 0

    # Общая статистика КЧС
    kchz_total = len(kchz_data)
    kchz_green = sum(1 for k in kchz_data if k.get('percentage', 0) >= 90)
    kchz_yellow = sum(1 for k in kchz_data if 70 <= k.get('percentage', 0) < 90)
    kchz_red = sum(1 for k in kchz_data if k.get('percentage', 0) < 70)

    return jsonify({
        'period': {
            'start': start_date.strftime('%d.%m.%Y'),
            'end': end_date.strftime('%d.%m.%Y'),
            'type': period
        },
        'summary': {
            'total': total,
            'completed': completed,
            'overdue': overdue,
            'in_progress': in_progress,
            'percentage': pct
        },
        'monthly': monthly_data,
        'top_executors': top_executors,
        'kchz': {
            'districts': kchz_data,
            'stats': {
                'total': kchz_total,
                'green': kchz_green,
                'yellow': kchz_yellow,
                'red': kchz_red
            }
        }
    })


# ─── API: Светофор КЧС (данные из Excel) ──────────────────────────────────────
@app.route('/api/kchz-traffic-light')
def api_kchz_traffic_light():
    """Чтение данных из файла светофоры.xlsx и возврат в формате JSON"""
    try:
        import openpyxl
        # Файл может быть в корне проекта (рядом с папкой app/) или в рабочей директории
        xlsx_path = os.path.join(BASE_DIR, '..', 'светофоры.xlsx')
        xlsx_path = os.path.abspath(xlsx_path)
        
        # Если не нашли — пробуем в текущей рабочей директории (Docker)
        if not os.path.exists(xlsx_path):
            xlsx_path = os.path.join(os.getcwd(), 'светофоры.xlsx')
        
        # Если всё ещё не нашли — пробуем рядом с main.py
        if not os.path.exists(xlsx_path):
            xlsx_path = os.path.join(BASE_DIR, 'светофоры.xlsx')
        
        if not os.path.exists(xlsx_path):
            return jsonify({"error": "Файл светофоры.xlsx не найден"}), 404
        
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        districts = []
        # Данные начинаются с строки 6 (индекс 5), столбцы: A=№, B=название, C=всего, D=выполнено, E=процент
        for row in range(6, ws.max_row + 1):
            num = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            total = ws.cell(row=row, column=3).value
            completed = ws.cell(row=row, column=4).value
            percentage = ws.cell(row=row, column=5).value
            
            # Пропускаем пустые строки и заголовки
            if not isinstance(num, (int, float)) or not name:
                continue
            
            # Определяем статус светофора
            if percentage is None or percentage == 0:
                color = 'red'
            elif percentage >= 0.9:
                color = 'green'
            elif percentage >= 0.7:
                color = 'yellow'
            else:
                color = 'red'
            
            districts.append({
                'num': int(num),
                'name': str(name).strip(),
                'total': int(total) if total else 0,
                'completed': int(completed) if completed else 0,
                'percentage': round(float(percentage) * 100, 1) if percentage else 0,
                'color': color,
            })
        
        wb.close()
        
        # Сортировка: зелёные → жёлтые → красные
        color_order = {'green': 0, 'yellow': 1, 'red': 2}
        districts.sort(key=lambda x: (color_order.get(x['color'], 2), -x['percentage']))
        
        # Статистика
        total_districts = len(districts)
        green_count = sum(1 for d in districts if d['color'] == 'green')
        yellow_count = sum(1 for d in districts if d['color'] == 'yellow')
        red_count = sum(1 for d in districts if d['color'] == 'red')
        
        return jsonify({
            'districts': districts,
            'stats': {
                'total': total_districts,
                'green': green_count,
                'yellow': yellow_count,
                'red': red_count,
            }
        })
        
    except Exception as e:
        logging.error(f"Ошибка чтения светофоры.xlsx: {e}")
        return jsonify({"error": str(e)}), 500


# ─── Загрузка светофоры.xlsx ─────────────────────────────────────────────────
@app.route('/api/kchz/upload', methods=['POST'])
def upload_kchz_file():
    """Загрузка файла светофоры.xlsx"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Файл не найден"}), 400
        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({"error": "Файл не выбран"}), 400
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Только .xlsx файлы"}), 400

        # Сохраняем в несколько возможных локаций
        xlsx_data = file.read()
        targets = [
            os.path.join(BASE_DIR, '..', 'светофоры.xlsx'),
            os.path.join(os.getcwd(), 'светофоры.xlsx'),
            os.path.join(BASE_DIR, 'светофоры.xlsx'),
        ]
        saved = False
        for target in targets:
            try:
                target = os.path.abspath(target)
                with open(target, 'wb') as f:
                    f.write(xlsx_data)
                saved = True
                logging.info(f"светофоры.xlsx сохранён в {target}")
            except Exception as e:
                logging.warning(f"Не удалось сохранить в {target}: {e}")

        if not saved:
            return jsonify({"error": "Не удалось сохранить файл"}), 500

        return jsonify({"ok": True, "message": "Файл светофоры.xlsx обновлён"})
    except Exception as e:
        logging.error(f"Ошибка загрузки светофоры.xlsx: {e}")
        return jsonify({"error": str(e)}), 500


# ─── Страница «Светофор КЧС» ─────────────────────────────────────────────────
@app.route('/kchz-traffic-light')
def kchz_traffic_light():
    """Страница визуализации светофора КЧС"""
    return render_template('kchz_traffic_light.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
